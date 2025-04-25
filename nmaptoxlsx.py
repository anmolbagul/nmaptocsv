import os
import xml.etree.ElementTree as ET
import csv
from collections import defaultdict
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Define extended service-port mappings
SERVICE_CATEGORIES = {
    'web': [80, 8080, 8000, 443, 8443],
    'ssh': [22],
    'ftp': [21],
    'rdp': [3389],
    'dns': [53],
    'smtp': [25, 465, 587],
    'pop3': [110, 995],
    'imap': [143, 993],
    'db': [1433, 1521, 3306, 5432, 6379, 9200, 27017],
    'file-sharing': [139, 445],
    'vpn': [1194, 500]
}

INPUT_DIR = './nmap_xml/'
OUTPUT_ALL_HOSTS = './output/all_hosts.xlsx'
OUTPUT_DETAILED = './output/port_detailed_report.xlsx'
os.makedirs(os.path.dirname(OUTPUT_ALL_HOSTS), exist_ok=True)

# Risk assessment based on script output
def assess_risk(script_output):
    output = script_output.lower()
    if 'vulnerable' in output or 'cve-' in output:
        return 'High'
    elif 'weak' in output or 'outdated' in output:
        return 'Medium'
    else:
        return 'Low'

# Extract CVE IDs from script output
def extract_cves(script_output):
    return ';'.join(re.findall(r'CVE-\d{4}-\d{4,7}', script_output))

# Categorize port to service bucket
def get_service_category(port, service_name):
    for category, ports in SERVICE_CATEGORIES.items():
        if port in ports:
            return category
    if service_name:
        name = service_name.lower()
        if 'http' in name or 'web' in name:
            return 'web'
        elif 'ssh' in name:
            return 'ssh'
        elif 'ftp' in name:
            return 'ftp'
        elif 'rdp' in name:
            return 'rdp'
        elif 'smtp' in name or 'mail' in name:
            return 'smtp'
        elif 'pop3' in name:
            return 'pop3'
        elif 'imap' in name:
            return 'imap'
        elif 'mysql' in name or 'mssql' in name or 'postgres' in name or 'mongo' in name or 'redis' in name:
            return 'db'
        elif 'smb' in name or 'netbios' in name:
            return 'file-sharing'
        elif 'vpn' in name:
            return 'vpn'
    return 'misc'

host_data = defaultdict(lambda: {
    'hostnames': set(),
    'os': 'unknown',
    'ip': '',
    'file': '',
    'timestamp': '',
    'ports': []
})

port_detailed_data = []

for filename in os.listdir(INPUT_DIR):
    if filename.endswith('.xml'):
        filepath = os.path.join(INPUT_DIR, filename)
        try:
            tree = ET.parse(filepath)
            root = tree.getroot()
            scan_time = datetime.fromtimestamp(os.path.getmtime(filepath)).isoformat()

            for host in root.findall('host'):
                address_el = host.find('address')
                address = address_el.attrib.get('addr') if address_el is not None else 'unknown'

                hostnames = []
                for hn in host.findall("hostnames/hostname"):
                    hostnames.append(hn.attrib.get('name'))

                os_guess = 'unknown'
                os_el = host.find("os")
                if os_el is not None:
                    best_os = os_el.find("osmatch")
                    if best_os is not None:
                        os_guess = best_os.attrib.get('name')

                ports = host.find('ports')
                if ports is None:
                    continue

                for port in ports.findall('port'):
                    state = port.find('state').attrib.get('state')
                    if state != 'open':
                        continue

                    proto = port.attrib.get('protocol')
                    portid = int(port.attrib.get('portid'))
                    service_el = port.find('service')
                    service_name = service_el.attrib.get('name') if service_el is not None else 'unknown'
                    product = service_el.attrib.get('product') if service_el is not None and 'product' in service_el.attrib else ''
                    service_fp = service_el.attrib.get('ostype') if service_el is not None and 'ostype' in service_el.attrib else ''

                    category = get_service_category(portid, service_name)

                    script_id = ''
                    script_output = ''
                    risk = 'Low'
                    cve_list = ''

                    for script in port.findall('script'):
                        script_id = script.attrib.get('id', '')
                        script_output = script.attrib.get('output', '')
                        risk = assess_risk(script_output)
                        cve_list = extract_cves(script_output)

                    port_detailed_data.append([
                        portid,
                        address,
                        ';'.join(hostnames) if hostnames else 'N/A',
                        os_guess,
                        proto,
                        service_name,
                        category,
                        product,
                        service_fp,
                        script_id,
                        script_output,
                        risk,
                        cve_list,
                        filename,
                        ''  # Notes
                    ])

                    host_data[address]['hostnames'].update(hostnames)
                    host_data[address]['os'] = os_guess
                    host_data[address]['ip'] = address
                    host_data[address]['file'] = filename
                    host_data[address]['timestamp'] = scan_time
                    host_data[address]['ports'].append((portid, proto, service_name, category, product, service_fp, script_id, script_output, risk, cve_list))

        except ET.ParseError:
            print(f"[!] Failed to parse {filename}")

# Sort port-detailed data by Port then IP
port_detailed_data.sort(key=lambda x: (x[0], x[1]))

# Write port-detailed Excel
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Port Details"
header1 = ['Port', 'IP', 'Host', 'OS', 'Protocol', 'Service', 'Category', 'Product', 'Service FP', 'NSE Script ID', 'NSE Script Output', 'Risk Level', 'CVE List', 'Source File', 'Notes']
ws1.append(header1)
for row in port_detailed_data:
    ws1.append(row)
ws1.auto_filter.ref = f"A1:{get_column_letter(len(header1))}{len(port_detailed_data)+1}"
wb1.save(OUTPUT_DETAILED)

# Write all-hosts summary Excel
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "All Hosts"
host_header = ['IP', 'Hostnames', 'OS', 'Port Count', 'Ports', 'Protocols', 'Services', 'Categories', 'Products', 'Service FPs', 'NSE Script IDs', 'NSE Script Outputs', 'Risk Levels', 'CVE Lists', 'Source File', 'Timestamp']
ws2.append(host_header)

for ip, data in host_data.items():
    ports = [str(p[0]) for p in data['ports']]
    protocols = [p[1] for p in data['ports']]
    services = [p[2] for p in data['ports']]
    categories = [p[3] for p in data['ports']]
    products = [p[4] for p in data['ports']]
    service_fps = [p[5] for p in data['ports']]
    script_ids = [p[6] for p in data['ports']]
    script_outputs = [p[7] for p in data['ports']]
    risks = [p[8] for p in data['ports']]
    cves = [p[9] for p in data['ports']]
    ws2.append([
        data['ip'],
        ';'.join(data['hostnames']) if data['hostnames'] else 'N/A',
        data['os'],
        len(data['ports']),
        ';'.join(ports),
        ';'.join(protocols),
        ';'.join(services),
        ';'.join(categories),
        ';'.join(products),
        ';'.join(service_fps),
        ';'.join(script_ids),
        ';'.join(script_outputs),
        ';'.join(risks),
        ';'.join(cves),
        data['file'],
        data['timestamp']
    ])

ws2.auto_filter.ref = f"A1:{get_column_letter(len(host_header))}{ws2.max_row}"
wb2.save(OUTPUT_ALL_HOSTS)

print(f"[+] Host summary with vulnerability detection written to {OUTPUT_ALL_HOSTS}")
print(f"[+] Port details with vulnerability detection written to {OUTPUT_DETAILED}")
